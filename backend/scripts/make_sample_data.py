"""Generate a realistic STAND-IN data pack for local testing and demos.

This is NOT the official assessment pack — it mirrors the described shape (6 PDFs + a
workbook with account/order/ticket data and a README snapshot time) so the full pipeline
can be exercised end-to-end before the real pack is dropped into ../data.

Run:  python scripts/make_sample_data.py
Output: ../sample_data/  (point DATA_DIR there, or copy into ../data)

The content is intentionally seeded with conflicts and a deprecated doc so the
reliability/precedence logic has something real to resolve.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

OUT = Path(__file__).resolve().parents[1].parent / "sample_data"
OUT.mkdir(exist_ok=True)

SNAPSHOT = dt.datetime(2025, 8, 21, 9, 0, 0)


def pdf(name: str, title: str, paragraphs: list[str]) -> None:
    doc = SimpleDocTemplate(str(OUT / name), pagesize=LETTER)
    styles = getSampleStyleSheet()
    flow = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    for p in paragraphs:
        if p.startswith("## "):
            flow.append(Paragraph(p[3:], styles["Heading2"]))
        else:
            flow.append(Paragraph(p, styles["BodyText"]))
        flow.append(Spacer(1, 6))
    doc.build(flow)
    print("wrote", name)


pdf(
    "01_Support_Policy_v3_CURRENT.pdf",
    "ParcelPilot Support Policy v3 (CURRENT)",
    [
        "Effective 2025-06-01. This is the current general support policy and supersedes v2.",
        "## Cancellations",
        "Standard and Business customers may cancel an order free of charge up to 24 hours "
        "before the scheduled pickup time. Cancellations made within 24 hours of pickup incur "
        "a cancellation fee of 10% of the order value.",
        "Enterprise customers are governed by their individual enterprise agreement, which "
        "overrides this general policy.",
        "## Service Credits",
        "If ParcelPilot or its carrier partner is at fault for a service failure (for example a "
        "late pickup caused by carrier fault), the customer is entitled to a service credit of "
        "10% of the affected order value. Credits above the agent approval limit require manager "
        "approval.",
        "## Support SLAs",
        "Standard: first response within 24 hours. Business: within 8 hours. Enterprise: within "
        "4 hours, per agreement.",
    ],
)

pdf(
    "02_Support_Policy_v2_DEPRECATED.pdf",
    "ParcelPilot Support Policy v2 (DEPRECATED)",
    [
        "Effective 2024-01-01. DEPRECATED and superseded by v3 on 2025-06-01. Retained for "
        "historical reference only.",
        "## Cancellations",
        "All customers may cancel free of charge up to 48 hours before pickup; otherwise a 5% "
        "cancellation fee applies.",
        "## Service Credits",
        "Service credits for carrier-fault failures are 5% of the order value.",
    ],
)

pdf(
    "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
    "Cancellation & Service Credit SOP v4",
    [
        "Current operating procedure for agents handling cancellations and service credits.",
        "## Determining a cancellation fee",
        "1. Identify the order and its account. 2. If the account is Enterprise, apply the terms "
        "of that account's enterprise agreement. 3. Otherwise apply Support Policy v3: free "
        "cancellation up to 24 hours before pickup, else a 10% fee.",
        "## Service credit for late pickup (carrier fault)",
        "A pickup delayed by more than 2 hours due to carrier fault qualifies for a service "
        "credit. For non-enterprise accounts the credit is 10% of the order value per Policy v3. "
        "For enterprise accounts apply the agreement's credit rate. Credits exceeding the agent's "
        "approval limit must be escalated for manager approval.",
    ],
)

pdf(
    "04_Product_Operations_Guide_and_Known_Issues.pdf",
    "Product Operations Guide & Known Issues",
    [
        "## Known issue: CarrierX pickup delays",
        "CarrierX has intermittent pickup scheduling delays in the western region as of "
        "2025-08. Delays attributable to CarrierX are treated as carrier fault for service-credit "
        "purposes.",
        "## Tracking updates",
        "Tracking events may lag by up to 30 minutes; this alone is not a service failure.",
    ],
)

pdf(
    "05_Northstar_Logistics_Enterprise_Agreement.pdf",
    "Northstar Logistics — Enterprise Agreement",
    [
        "This agreement governs the Northstar Logistics account and overrides ParcelPilot's "
        "general support policy where they conflict.",
        "## Cancellations",
        "Northstar Logistics may cancel any order free of charge up to 2 hours before the "
        "scheduled pickup time. No cancellation fee applies within that window.",
        "## Service Credits",
        "For service failures caused by carrier fault, Northstar receives a service credit of 15% "
        "of the affected order value (higher than the general 10%).",
        "## Support SLA",
        "First response within 4 hours, 24/7.",
    ],
)

pdf(
    "06_LumenWorks_Service_Agreement.pdf",
    "LumenWorks — Service Agreement",
    [
        "This agreement governs the LumenWorks account.",
        "## Cancellations",
        "LumenWorks may cancel free of charge up to 24 hours before pickup. One late "
        "cancellation fee waiver is permitted per calendar quarter.",
        "## Service Credits",
        "Carrier-fault service failures receive the standard 10% service credit.",
    ],
)

# --- Workbook --------------------------------------------------------------
wb = Workbook()

readme = wb.active
readme.title = "README"
readme.append(["ParcelPilot Assessment Data (SAMPLE)"])
readme.append(["Dataset snapshot time", SNAPSHOT.strftime("%Y-%m-%d %H:%M:%S")])
readme.append(["Note", "Use the snapshot time above as 'now' for all time-based questions."])

accounts = wb.create_sheet("Accounts")
accounts.append(["account_id", "account_name", "tier", "contact_email", "contact_phone",
                 "contract_value"])
for row in [
    ["ACC-100", "Northstar Logistics", "Enterprise", "ops@northstar.example", "+91-9800000001", 5000000],
    ["ACC-200", "LumenWorks", "Business", "help@lumenworks.example", "+91-9800000002", 900000],
    ["ACC-300", "Bright Retail", "Standard", "care@brightretail.example", "+91-9800000003", 120000],
]:
    accounts.append(row)

orders = wb.create_sheet("Orders")
orders.append(["order_id", "account_id", "account_name", "pickup_scheduled", "status",
               "order_value", "carrier", "service"])
for row in [
    ["ORD-1001", "ACC-100", "Northstar Logistics", "2025-08-22 14:00:00", "Booked", 12000, "CarrierX", "express"],
    ["ORD-1002", "ACC-200", "LumenWorks", "2025-08-21 18:00:00", "Booked", 4000, "CarrierY", "standard"],
    ["ORD-1003", "ACC-300", "Bright Retail", "2025-08-21 12:00:00", "In Transit", 1500, "CarrierX", "standard"],
    ["ORD-1004", "ACC-100", "Northstar Logistics", "2025-08-20 09:00:00", "Delivered", 8000, "CarrierZ", "express"],
]:
    orders.append(row)

tickets = wb.create_sheet("Tickets")
tickets.append(["ticket_id", "account_id", "account_name", "created_at", "severity", "status",
                "category", "sla_hours", "subject", "resolution_note"])
for row in [
    ["TKT-5001", "ACC-100", "Northstar Logistics", "2025-08-19 08:00:00", "high", "Open", "pickup_delay", 4,
     "Pickup 3h late for ORD-1004", ""],
    ["TKT-5002", "ACC-300", "Bright Retail", "2025-08-20 22:00:00", "medium", "Open", "pickup_delay", 24,
     "CarrierX pickup delayed", ""],
    ["TKT-5003", "ACC-200", "LumenWorks", "2025-08-21 06:00:00", "low", "Open", "pickup_delay", 24,
     "Late pickup on standard order", ""],
    ["TKT-5004", "ACC-100", "Northstar Logistics", "2025-08-18 10:00:00", "high", "Open", "billing", 4,
     "Disputed cancellation fee", "Told customer 5% fee applies"],
    ["TKT-5005", "ACC-200", "LumenWorks", "2025-08-15 09:00:00", "medium", "Resolved", "tracking", 24,
     "Tracking not updating", "Explained 30-min lag; resolved"],
    ["TKT-5006", "ACC-100", "Northstar Logistics", "2025-08-21 07:30:00", "high", "Open", "pickup_delay", 4,
     "Repeat pickup delay CarrierX", ""],
]:
    tickets.append(row)

xlsx_path = OUT / "ParcelPilot_Assessment_Data.xlsx"
wb.save(str(xlsx_path))
print("wrote", xlsx_path.name)
print("\nSample pack written to", OUT)

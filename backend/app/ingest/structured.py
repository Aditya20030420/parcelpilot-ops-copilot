"""Structured-data ingestion for ParcelPilot_Assessment_Data.xlsx.

The workbook holds account, order, and ticket data across several sheets. We do NOT know
the exact column names ahead of time, so this loader introspects each sheet: the first
non-empty row is treated as the header, remaining rows become dict records.

We deliberately expose a *parameterised* query API (query_table with whitelisted
operators) rather than letting the model write raw SQL. That keeps the data/tool layer in
control of what can be read, which is what makes account/role scoping enforceable.

The dataset "snapshot time" is read from the README sheet and used as the reference
`now` for every time-based question, per the assessment instructions.
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _norm(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def _parse_dt(v: Any) -> dt.datetime | None:
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M",
                    "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return dt.datetime.strptime(s, fmt)
            except ValueError:
                continue
        # ISO with timezone
        try:
            return dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


@dataclass
class Table:
    name: str
    columns: list[str]
    rows: list[dict[str, Any]] = field(default_factory=list)


_OPERATORS = {
    "eq": lambda a, b: _cmp_eq(a, b),
    "neq": lambda a, b: not _cmp_eq(a, b),
    "contains": lambda a, b: b is not None and str(b).lower() in str(a).lower(),
    "gt": lambda a, b: _num(a) is not None and _num(b) is not None and _num(a) > _num(b),
    "gte": lambda a, b: _num(a) is not None and _num(b) is not None and _num(a) >= _num(b),
    "lt": lambda a, b: _num(a) is not None and _num(b) is not None and _num(a) < _num(b),
    "lte": lambda a, b: _num(a) is not None and _num(b) is not None and _num(a) <= _num(b),
    "in": lambda a, b: str(a).lower() in [str(x).lower() for x in (b or [])],
}


def _cmp_eq(a: Any, b: Any) -> bool:
    if a is None or b is None:
        return a == b
    return str(a).strip().lower() == str(b).strip().lower()


def _num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        m = re.search(r"-?\d+(\.\d+)?", v.replace(",", ""))
        if m:
            return float(m.group())
    return None


class StructuredStore:
    def __init__(self) -> None:
        self.tables: dict[str, Table] = {}
        self.snapshot_time: dt.datetime | None = None
        self.readme_text: str = ""

    def load_file(self, path: Path) -> None:
        wb = load_workbook(str(path), data_only=True, read_only=True)
        for ws in wb.worksheets:
            name = ws.title.strip()
            if name.lower() in {"readme", "read me", "read_me", "notes"}:
                self._load_readme(ws)
                continue
            self._load_sheet(ws)
        wb.close()

    def _load_readme(self, ws) -> None:
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
        self.readme_text = "\n".join(lines)
        # Find a snapshot / reference / as-of timestamp anywhere in the README.
        for line in lines:
            if re.search(r"snapshot|as of|as-of|reference time|current time|dataset time", line, re.I):
                found = self._first_dt_in(line)
                if found:
                    self.snapshot_time = found
                    return
        # Fallback: first parseable datetime anywhere in the README.
        for line in lines:
            found = self._first_dt_in(line)
            if found:
                self.snapshot_time = found
                return

    @staticmethod
    def _first_dt_in(line: str) -> dt.datetime | None:
        for m in re.finditer(
            r"\d{4}-\d{2}-\d{2}(?:[ T]\d{2}:\d{2}(?::\d{2})?)?|\d{1,2}/\d{1,2}/\d{4}", line
        ):
            parsed = _parse_dt(m.group())
            if parsed:
                return parsed
        return None

    def _load_sheet(self, ws) -> None:
        rows_iter = ws.iter_rows(values_only=True)
        header: list[str] | None = None
        records: list[dict[str, Any]] = []
        for raw in rows_iter:
            values = [_norm(c) for c in raw]
            if header is None:
                if all(c is None or c == "" for c in values):
                    continue
                header = [str(c).strip() if c is not None else f"col{i}"
                          for i, c in enumerate(values)]
                continue
            if all(c is None or c == "" for c in values):
                continue
            rec = {header[i]: (values[i] if i < len(values) else None)
                   for i in range(len(header))}
            records.append(rec)
        if header:
            self.tables[self._canonical(ws.title)] = Table(
                name=ws.title.strip(), columns=header, rows=records
            )

    @staticmethod
    def _canonical(name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_")

    # --- Query API ----------------------------------------------------------
    def list_tables(self) -> list[dict]:
        return [{"table": key, "sheet_name": t.name, "columns": t.columns,
                 "row_count": len(t.rows)} for key, t in self.tables.items()]

    def resolve_table(self, name: str) -> Table | None:
        key = self._canonical(name)
        if key in self.tables:
            return self.tables[key]
        # loose contains-match, e.g. "orders" -> "order_data"
        for k, t in self.tables.items():
            if key in k or k in key:
                return t
        return None

    def query_table(
        self,
        table: str,
        filters: list[dict] | None = None,
        columns: list[str] | None = None,
        limit: int = 50,
    ) -> dict:
        t = self.resolve_table(table)
        if t is None:
            return {"error": f"Unknown table '{table}'. Available: "
                             f"{[x['table'] for x in self.list_tables()]}"}
        out: list[dict] = []
        for rec in t.rows:
            if self._matches(rec, filters or []):
                out.append(rec)
        selected = out
        if columns:
            keep = [c for c in columns if c in t.columns]
            selected = [{c: r.get(c) for c in keep} for r in out]
        return {
            "table": self._canonical(t.name),
            "matched": len(out),
            "returned": min(len(selected), limit),
            "rows": _jsonable(selected[:limit]),
        }

    def _matches(self, rec: dict, filters: list[dict]) -> bool:
        for f in filters:
            col, op, val = f.get("column"), f.get("op", "eq"), f.get("value")
            if col not in rec:
                return False
            fn = _OPERATORS.get(op)
            if fn is None:
                return False
            try:
                if not fn(rec.get(col), val):
                    return False
            except Exception:  # noqa: BLE001
                return False
        return True

    def stats(self) -> dict:
        return {
            "snapshot_time": self.snapshot_time.isoformat() if self.snapshot_time else None,
            "tables": self.list_tables(),
        }


def _jsonable(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            if isinstance(v, (dt.datetime, dt.date)):
                nr[k] = v.isoformat()
            else:
                nr[k] = v
        out.append(nr)
    return out
